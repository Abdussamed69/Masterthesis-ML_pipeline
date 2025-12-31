#2025-09-26 Author:             Abdussamed Korkmaz 
#                               Master Electrical Systems Engineering 
#                               University of Heilbronn, Germany
#2025-09-26 Description:        Script for working with excel files 
#************************************************************************************************************************************

import pandas as pd
# 'xlsxwriter' is required for writing excel files with pandas
from pandas import ExcelWriter
from openpyxl import load_workbook, Workbook
from pathlib import Path
from typing import Union, Optional

def excel_param_correct(input_dir: Union[str, Path], output_dir: Optional[Union[str, Path]] = None):
    """Korrigiert einen bestimmten Wert in allen Excel-Dateien in einem Verzeichnis."""
    input_dir = Path(input_dir)
    if output_dir:
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
    else:
        output_dir = input_dir  # Wenn kein Output-Ordner: überschreibe Dateien im Input-Ordner

    # Alle Excel-Dateien lesen
    file_paths = sorted(input_dir.glob("*.xlsx"))
    if not file_paths:
        raise FileNotFoundError("Keine Excel-Dateien im angegebenen Verzeichnis gefunden.")

    for file_path in file_paths:
        try:
            data = pd.read_excel(file_path, header=None)

            # Wert korrigieren
            data.loc[4, 2] = 300

            # Speichern unter gleichem Dateinamen im output_dir
            output_file = output_dir / file_path.name
            with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
                data.to_excel(writer, index=False, header=False)

            print(f"Korrigiert: {file_path.name}")

        except Exception as e:
            print(f"Fehler bei Datei {file_path.name}: {e}")

def excel_param_correct_withFormat(input_dir: Union[str, Path],
                                   output_dir: Optional[Union[str, Path]] = None,
                                   Row: int = 3,
                                   Column: int = 2,
                                   Value: str = "data_"):
    """Korrigiert einen bestimmten Wert in allen Excel-Dateien in einem Verzeichnis mit Formatierung."""
    input_dir = Path(input_dir)
    if output_dir:
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
    else:
        output_dir = input_dir  # Wenn kein Output-Ordner: überschreibe Dateien im Input-Ordner

    # load .xlsx-Data
    file_paths = list(input_dir.glob("*.xlsx"))
    if not file_paths:
        raise FileNotFoundError("No .xlsx-file found in folder!")

    for file_path in file_paths:
        try:
            wb = load_workbook(file_path)
            ws = wb["Info"]

            # Cell e.g. C3 or A1 according to Excel columns-rows (row, column)
            ws.cell(row=Row, column=Column).value = Value  # Change value in cell 

            # save with same file name as in <output_dir> folder
            save_path = output_dir / file_path.name
            wb.save(save_path)

            print(f"data file is corrected & saved: {save_path.name}")

        except Exception as e:
            print(f"Error in data file {file_path.name}: {e}")

def concat_excels(input_dir: Union[str, Path], output_dir: Optional[Union[str, Path]] = None, 
                  output_file: str = "classification_matrix.xlsx"):
    """
    Fügt alle Matrizen (Excel-Dateien) aus einem Ordner zu einer großen Matrix zusammen (zeilenweise).

    :param input_dir: Pfad zum Ordner mit .xlsx-Dateien
    :param output_dir: Ziel-Ordner. Wenn None -> input_dir / output_file
    :param output_file: Dateiname
    """
    input_dir = Path(input_dir)
    if not input_dir.exists():
        raise FileNotFoundError(f"Input directory does not exist: {input_dir}")
    
    if output_dir:
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        output = output_dir / output_file
    else:
        output_dir = input_dir  # Wenn kein Output-Ordner: überschreibe Dateien im Input-Ordner
        output = output_dir / output_file

    # Liste .xlsx Dateien, ignoriere Excel-Tempfiles (~$...)
    file_paths = sorted(p for p in input_dir.glob("*.xlsx") if not p.name.startswith("~$"))
    if not file_paths:
        raise FileNotFoundError("Keine Excel-Dateien im Verzeichnis gefunden.")

    all_dfs = []
    for file_path in file_paths:
        try:
            df = pd.read_excel(file_path, engine='openpyxl')
            all_dfs.append(df)
            print(f"Gelesen: {file_path.name}")
        except Exception as e:
            print(f"Fehler beim Lesen von {file_path.name}: {e}")

    if not all_dfs:
        raise ValueError("Keine Dateien konnten gelesen werden.")

    # Zeilenweise zusammenfügen
    merged_df = pd.concat(all_dfs, axis=0, ignore_index=True)

    # entferne Zeilen mit in exclude definierten OilTypes
    # exclude = Ölklassen, die ausgeschlossen werden sollen
    if "OilType" in merged_df.columns:
            #exclude = {"reference", "extra virgin olive oil", "refined olive-pomace oil", "olive-pomace oil", "mixture sunflower-olive oil 90-10"}  # Referenz und reine Öle ausschließen
            #exclude = {"reference", "blend", "mixture sunflower-olive oil 90-10"}  # Mischungen und Referenz ausschließen
            exclude = {"reference", "mixture sunflower-olive oil 90-10"}
            mask = ~merged_df["OilType"].isin(exclude)
            merged_df = merged_df.loc[mask].reset_index(drop=True)

    # Speichern
    merged_df.to_excel(output, index=False)
    print(f"Zusammenfuehrung abgeschlossen. Gespeichert unter: {output}")

def rename_excelFiles(input_dir: Union[str, Path], newname: str = "data_"):
    """This function renames all excel files in a directory to a new name (all same name)"""
    input_dir = Path(input_dir)
    file_paths = list(input_dir.glob("*.xlsx"))
    if not file_paths:
        raise FileNotFoundError("Keine Excel-Dateien im Verzeichnis gefunden.")

    for i, file_path in enumerate(file_paths, start=1):
        new_name = f"{newname}.xlsx"
        new_path = input_dir / new_name
        file_path.rename(new_path)
        print(f"Umbenannt: {file_path.name} -> {new_name}")

def replace_String_excelFiles(input_dir: Union[str, Path], 
                                 replaceString: str = "data_1", 
                                 newString: str = "data_2"):
    
    """This function removes a specific part of the filename from all excel files in a directory"""
    input_dir = Path(input_dir)
    file_paths = list(input_dir.glob("*.xlsx"))
    if not file_paths:
        raise FileNotFoundError("Keine Excel-Dateien im Verzeichnis gefunden.")

    for i, file_path in enumerate(file_paths, start=1):
        # Wenn der zu ersetzende String im Dateinamen ist
        if replaceString in file_path.stem:
            original_stem = file_path.stem  # Dateiname ohne Erweiterung
            processed_stem = original_stem.replace(replaceString, newString) # ersetze das Pattern
            new_name = f"{processed_stem}.xlsx"
            new_path = input_dir / new_name
            file_path.rename(new_path)
            print(f"Umbenannt: {file_path.name} -> {new_name}")
        else:
            print(f"'{replaceString}' nicht in {file_path.name} gefunden. Keine Änderung vorgenommen.")

def replace_List_of_Strings_excelFiles(input_dir: Union[str, Path],
                             replaceList: Union[str, list, tuple, set] = "data_1",
                             newString: str = "data_2"):
    """
    Entfernt in Dateinamen alle Vorkommen von replaceString (oder allen Strings in einer Liste)
    und ersetzt sie durch newString.
    Beispiel: replaceString = ["meas1-", "meas2-", ...], newString = ""
    """
    input_dir = Path(input_dir)
    file_paths = list(input_dir.glob("*.xlsx"))
    if not file_paths:
        raise FileNotFoundError("Keine Excel-Dateien im Verzeichnis gefunden.")

    # Normalisiere replaceString zu einer Liste von Patterns
    if isinstance(replaceList, (list, tuple, set)):
        patterns = list(replaceList)
    else:
        patterns = [str(replaceList)]

    for i, file_path in enumerate(file_paths, start=1):
        if replaceList in file_path.stem:
            original_stem = file_path.stem  # Dateiname ohne Erweiterung
            new_stem = original_stem
            # entferne alle Pattern-Vorkommen
            for pat in patterns:
                new_stem = new_stem.replace(str(pat), newString)
            # optional: trim whitespace
            processed_stem = new_stem.strip()
            new_name = f"{processed_stem}.xlsx"
            new_path = input_dir / new_name
        else:
            print(f"Eines der Patterns {patterns} nicht in {file_path.name} gefunden. Keine Änderung vorgenommen.")
            continue

        # Kollision vermeiden: füge Zähler an wenn Name bereits existiert und nicht mit sich selbst übereinstimmt
        counter = 1
        base = processed_stem
        while new_path.exists() and new_path != file_path:
            new_name = f"{counter}-{base}.xlsx"
            new_path = input_dir / new_name
            counter += 1

        file_path.rename(new_path)
        print(f"Umbenannt: {file_path.name} -> {new_name}")

def excel_split_measurements(input_dir: Union[str, Path], output_dir: Optional[Union[str, Path]] = None):
    """diese Funktion splittet Excel-Dateien mit mehreren Messungen in einzelne Dateien auf.
    Jede neue Datei enthält nur eine Messung, und die 'Info'-Tabelle wird entsprechend aktualisiert."""
    input_dir = Path(input_dir)
    if output_dir:
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
    else:
        output_dir = input_dir
    print(f"Output directory: {output_dir}")

    # Get all Excel files in the input directory
    file_paths = list(input_dir.glob("*.xlsx"))
    if not file_paths:
        raise FileNotFoundError("No .xlsx files found in the directory.")

    for file_path in file_paths:
        try:
            wb = load_workbook(file_path)
            if "Info" not in wb.sheetnames:
                print(f"'Info' sheet is missing in {file_path.name}, skipping file.")
                continue

            ws_info = wb["Info"]
            nmeas = ws_info.cell(row=6, column=3).value  # Cell C6 contains number of measurements

            if not isinstance(nmeas, int) or nmeas < 1:
                print(f"Invalid measurement count ({nmeas}) in file: {file_path.name}")
                continue
            if nmeas == 1:
                print(f"{file_path.name}: Only one measurement – no split needed.")
                continue

            print(f"\nSplitting {file_path.name}: {nmeas} measurements...")

            ws_vis = wb["VIS"] if "VIS" in wb.sheetnames else None
            ws_nir = wb["NIR"] if "NIR" in wb.sheetnames else None

            # Extract wavelength columns from VIS and NIR (column B)
            wavelengths_vis = [cell.value for cell in ws_vis["B"][1:]] if ws_vis else []
            wavelengths_nir = [cell.value for cell in ws_nir["B"][1:]] if ws_nir else []

            for i in range(nmeas):
                new_wb = Workbook()
                new_wb.remove(new_wb.active)  # Remove default sheet

                # Create VIS sheet
                if ws_vis:
                    vis_ws = new_wb.create_sheet("VIS")
                    vis_ws.append(["Index", "Wavelength", "Counts"])
                    for j, wl in enumerate(wavelengths_vis):
                        val = ws_vis.cell(row=j+2, column=3+i).value  # Measurements start at col 3 (C)
                        vis_ws.append([j+1, wl, val])

                # Create NIR sheet
                if ws_nir:
                    nir_ws = new_wb.create_sheet("NIR")
                    nir_ws.append(["Index", "Wavelength", "Counts"])
                    for j, wl in enumerate(wavelengths_nir):
                        val = ws_nir.cell(row=j+2, column=3+i).value
                        nir_ws.append([j+1, wl, val])

                # Create Info sheet and update measurement count to 1
                info_ws = new_wb.create_sheet("Info")
                for row in ws_info.iter_rows():
                    for cell in row:
                        info_ws[cell.coordinate].value = cell.value
                info_ws.cell(row=6, column=3).value = 1  # Set C6 = 1 (only one measurement now)

                # Save the new file
                new_filename = f"meas{i+1}-" + file_path.stem + f".xlsx"
                save_path = output_dir / new_filename
                new_wb.save(save_path)
                print(f"Saved: {save_path.name}")

        except Exception as e:
            print(f"Error processing {file_path.name}: {e}")             

def excel_average_values_of_multiple_files(input_dir: Union[str, Path], output_dir: Optional[Union[str, Path]] = None):
    """Diese Funktion berechnet den Durchschnittswert spaltenweise für jede Messung über mehrere Excel-Dateien hinweg
    und speichert die durchschnittlichen Werte in einer neuen Excel-Datei."""
    input_dir = Path(input_dir)
    if output_dir:
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
    else:
        output_dir = input_dir

    file_paths = list(input_dir.glob("*.xlsx"))
    if not file_paths:
        raise FileNotFoundError("No .xlsx files found in the directory.")

    all_dfs = []
    for file_path in file_paths:
        try:
            df = pd.read_excel(file_path, engine='openpyxl')
            all_dfs.append(df)
            print(f"Gelesen: {file_path.name}")
        except Exception as e:
            print(f"Fehler beim Lesen von {file_path.name}: {e}")

    if not all_dfs:
        raise ValueError("Keine Dateien konnten gelesen werden.")

    # Berechnung des Durchschnitts pro Spalte
    numeric_cols = all_dfs[0].select_dtypes(include='number').columns
    numeric_means = pd.concat([df[numeric_cols] for df in all_dfs]).mean(axis=0)

    # nicht-numerische Spalten vom ersten DataFrame übernehmen
    non_numeric_cols = all_dfs[0].select_dtypes(exclude='number').columns
    non_numeric_data = all_dfs[0][non_numeric_cols].iloc[0] #OilType and Source columns

    # Erstelle das Ergebnis-DataFrame
    # Kombiniere numerische Mittelwerte und nicht-numerische Daten
    result_df = pd.DataFrame([numeric_means])
    for col in non_numeric_cols:
        result_df[col] = non_numeric_data[col]

    # Speichern der Ergebnis-DataFrame in einer neuen Excel-Datei
    output_file = output_dir / "average_classification_matrix.xlsx"
    result_df.to_excel(output_file, index=False)
    print(f"Durchschnittswerte gespeichert unter: {output_file}")

def convert_csv_to_excel(input_dir: Union[str, Path], output_dir: Optional[Union[str, Path]] = None):
    """Konvertiert CSV-Dateien in Excel-Dateien."""
    input_dir = Path(input_dir)
    if not input_dir.exists():
        raise FileNotFoundError(f"Eingabeverzeichnis existiert nicht: {input_dir}")
    if output_dir:
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
    else:
        output_dir = input_dir  # Wenn kein Output-Ordner: überschreibe Dateien im Input-Ordner

    #csv_files = list(input_dir.glob("*.csv"))
    #if not csv_files:
    #    raise FileNotFoundError("Keine CSV-Dateien im angegebenen Verzeichnis gefunden.")
    
    # Loop über alle .csv Dateien
    for csv_file in input_dir.glob("*.csv"):
        try:
            df = pd.read_csv(csv_file)
            # Build output file path from csv filename
            out_filename = csv_file.with_suffix('.xlsx').name
            out_path = output_dir / out_filename
            # Write to an .xlsx file (pandas will choose a suitable engine)
            df.to_excel(out_path, index=False)
            print(f"Konvertiert: {csv_file.name} -> {out_path.name}")
        except Exception as e:
            print(f"Fehler bei der Konvertierung von {csv_file.name}: {e}")

if __name__ == "__main__":
    #***************Beispielaufrufe der Funktionen***************    
    #excel_param_correct
    #input_dir = Path(r"C:\Users\korkm\Desktop\5.Sem (Master-Thesis)\Thesis\Messungen\275nm\Fake oil 3_sunflower SUNTAT")
    #output_dir = Path(r"C:\Users\korkm\Desktop\Master\5.Sem (Master-Thesis)\Thesis\Messungen\lamp\Zeliva - refined olive pomace oil_corrected")
    #excel_param_correct_withFormat(input_dir=input_dir, Row=3, Column=3, Value="V2-Fake oil 3_out-of-class")

    #concat_excels
    input_dir = Path(r"C:\Users\korkm\Desktop\5.Sem (Master-Thesis)\Thesis\Messungen\lamp\lamp_Savitzky-Golay-Filter_allData\NIR")
    output_dir = Path(r"C:\Users\korkm\Desktop\5.Sem (Master-Thesis)\Thesis\Messungen\lamp\lamp_Savitzky-Golay-Filter_allData_matrix\NIR")
    output_file = "lamp_NIR_classification_matrix.xlsx"
    concat_excels(input_dir, output_dir, output_file)

    input_dir = Path(r"C:\Users\korkm\Desktop\5.Sem (Master-Thesis)\Thesis\Messungen\lamp\lamp_Savitzky-Golay-Filter_allData\VIS")
    output_dir = Path(r"C:\Users\korkm\Desktop\5.Sem (Master-Thesis)\Thesis\Messungen\lamp\lamp_Savitzky-Golay-Filter_allData_matrix\VIS")
    output_file = "lamp_VIS_classification_matrix.xlsx"
    concat_excels(input_dir, output_dir, output_file)

    #rename_excelFiles
    #input_dir = Path(r"C:\Users\korkm\Desktop\Master\5.Sem (Master-Thesis)\Thesis\Messungen\275nm\LIDL_rapeseed oil")
    #replace_String_excelFiles(input_dir = input_dir, 
    #                             replaceString = "Fake", 
    #                             newString = "V2-Fake")
    
    # split measurements
    #input_dir = Path(r"C:\Users\korkm\Desktop\5.Sem (Master-Thesis)\Thesis\Messungen\275nm\80 refined - 20 extra native\new")
    #output_dir = Path(r"C:\Users\korkm\Desktop\5.Sem (Master-Thesis)\Thesis\Messungen\275nm\Fake oil 3_sunflower SUNTAT\V2")
    #excel_split_measurements(input_dir, output_dir)
    
    # average values of multiple files
    #input_dir = Path(r"C:\Users\korkm\Desktop\Master\5.Sem (Master-Thesis)\Thesis\Messungen\lamp\Zeliva\NIR_Avg\allMeas")
    #output_dir = Path(r"C:\Users\korkm\Desktop\Master\5.Sem (Master-Thesis)\Thesis\Messungen\lamp\Zeliva\NIR_Avg\allMeas_Avg")
    #excel_average_values_of_multiple_files(input_dir, output_dir)

    # convert csv to excel
    #input_dir = Path(r"C:\Users\korkm\Documents\Abdussamed\Evlilik\Mervem\Bitpanda\Etherscan-CSV")
    #output_dir = Path(r"C:\Users\korkm\Documents\Abdussamed\Evlilik\Mervem\Bitpanda\Etherscan-CSV\Test")
    #convert_csv_to_excel(input_dir,output_dir)